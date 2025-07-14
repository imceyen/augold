from PIL import Image
import pytesseract
import re
import openpyxl
from openpyxl.utils import get_column_letter

# Tesseract 설치 경로
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# 이미지 경로
image_path = r"C:\ncsGlobal\FinalProject\augold\python\receipt\receipt_file\expense_20250714_163339.png"

# OCR 함수
def ocr_image_to_text(path):
    img = Image.open(path)
    text = pytesseract.image_to_string(img, lang='kor+eng')
    return text

def parse_receipt(text):
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    result = {
        "store": None,
        "total": None,
        "card_number": None,
        "transaction_time": None,
        "approval_number": None
    }

    for line in lines:
        m_store = re.search(r"가\s*맹\s*점\s*:\s*(.+)", line)
        if m_store:
            result["store"] = m_store.group(1).replace(" ", "")
            continue

        m_card = re.search(r"카\s*드\s*번\s*호\s*:\s*(.+)", line)
        if m_card:
            result["card_number"] = m_card.group(1).strip()
            continue

        m_approval = re.search(r"승\s*인\s*번\s*호\s*:\s*(\d+)", line)
        if m_approval:
            result["approval_number"] = m_approval.group(1)
            continue

        m_time = re.search(r"거\s*래\s*일\s*시\s*:\s*([\d/: ]+)", line)
        if m_time:
            result["transaction_time"] = m_time.group(1).strip()
            continue

        m_total = re.search(r"합\s*계\s*:\s*([\d,]+)\s*원", line)
        if m_total:
            result["total"] = m_total.group(1).replace(",", "")
            continue

    return result

def save_to_excel(data, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipt"

    # 헤더 작성
    headers = list(data.keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 데이터 작성 (1행에 헤더, 2행에 값)
    for col, key in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=data[key])

    # 열 너비 조정 (선택사항)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filepath)
    print(f"엑셀 파일이 저장되었습니다: {filepath}")

if __name__ == "__main__":
    text = ocr_image_to_text(image_path)
    print("OCR 결과 원문:\n" + "-"*40)
    print(text)
    print("-"*40)

    parsed = parse_receipt(text)
    print("🧾 파싱 결과:\n" + "-"*40)
    for k, v in parsed.items():
        print(f"{k}: {v}")

    # 엑셀 저장 경로 지정 (원하는 경로로 수정 가능)
    excel_path = r"C:\ncsGlobal\FinalProject\augold\python\receipt\parsed_receipt\parsed_receipt.xlsx"
    save_to_excel(parsed, excel_path)
